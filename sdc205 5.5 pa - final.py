print("holheb7286")

import os
import csv
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, PieChart
from datetime import datetime


# 1. AskUser Function
def askUser():
    # This loop collects 5 numbers from the user and accumulates their total.
    total = 0

    for counter in range(5):
        number = int(input("Please enter a number: "))
        total += number  # accumulate values properly

    print(f"The total of the numbers you entered is {total}.")
    return total


# 2. AskIncome Function
def askIncome():
    # This loop collects names + incomes and appends them to final.csv

    file_path = r"C:/PythonFiles/FinalExam/final.csv"

    with open(file_path, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)

        print("Please enter the name and annual income for 5 people.")

        for i in range(5):
            name = input(f"Enter name for person {i+1}: ")
            income = int(input(f"Enter annual income for {name}: "))

            # append each entry as a new row
            writer.writerow([name, income])

    print(f"Successfully added 5 entries to {file_path}")


# 3. Excel Pie Chart Function
def excelPie():
    # Load existing CSV into Excel-compatible format
    file_path = r"C:/PythonFiles/FinalExam/final.csv"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IncomeData"

    # Read CSV and write into workbook
    with open(file_path, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Convert income column to ints (VERY IMPORTANT for charts)
    for row in range(2, ws.max_row + 1):
        ws[f"B{row}"] = int(ws[f"B{row}"].value)

    # Labels are in Column A (1), Data in Column B (2)
    labels = Reference(ws, min_col=1, min_row=1, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)

    # Create pie chart
    chart = PieChart()
    chart.title = f"holheb7286 {datetime.now().strftime('%B %d, %Y')}"

    chart.add_data(data)
    chart.set_categories(labels)

    # Place chart in sheet
    ws.add_chart(chart, "D5")

    # Save file
    wb.save(r"C:/PythonFiles/FinalExam/final.xlsx")


# 4. Matplotlib Bar Chart Function
def verticalBar():
    file_path = r"C:/PythonFiles/FinalExam/final.csv"

    names = []
    incomes = []

    # This loop reads CSV data and separates it into two lists
    with open(file_path, 'r') as f:
        reader = csv.reader(f)

        for row in reader:
            names.append(row[0])
            incomes.append(int(row[1]))

    # Create bar chart
    plt.bar(names, incomes)

    plt.xlabel("Name")
    plt.ylabel("Income")
    plt.xticks(rotation=0)
    plt.title(f"holheb7286 {datetime.now().strftime('%B %d, %Y')}")

    plt.show()


# FUNCTION CALLS (for testing)
askUser()
askIncome()
excelPie()
verticalBar()
